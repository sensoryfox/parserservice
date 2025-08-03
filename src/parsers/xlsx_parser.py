from __future__ import annotations

from io import BytesIO
from uuid import UUID
from typing import List

from openpyxl import load_workbook

from ..models import Line, ParseResult
from .base import BaseParser


class XlsxParser(BaseParser):
    """Парсер Excel (лист/строки)."""

    async def parse(
        self,
        *,
        doc_id: UUID,          # не используется, но для единообразия
        file_content: BytesIO,
        parse_images: bool = True,  # изображений в .xlsx почти не бывает
    ) -> ParseResult:
        wb = load_workbook(filename=file_content, data_only=True)
        lines: List[Line] = []
        line_no = 0

        for sheet in wb.worksheets:
            # Заголовок листа
            lines.append(
                Line(
                    line_no=line_no,
                    sheet_name=sheet.title,
                    block_type="sheet_title",
                    content=f"## Sheet: {sheet.title}",
                )
            )
            line_no += 1

            for row in sheet.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                row_txt = " | ".join(cells)
                lines.append(
                    Line(
                        line_no=line_no,
                        sheet_name=sheet.title,
                        block_type="table",
                        content=row_txt,
                    )
                )
                line_no += 1

        return ParseResult(lines=lines, images=[])